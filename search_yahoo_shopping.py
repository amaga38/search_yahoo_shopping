import os
import argparse
import openpyxl

from yahoo_api import searchItems

from functools import wraps
import time

def stop_watch(func):
    @wraps(func)
    def wrapper(*args, **kargs):
        start = time.perf_counter()
        result = func(*args, **kargs)
        elapsed_time = time.perf_counter() - start
        print('Time: {} s ({} min)'.format(elapsed_time, elapsed_time//60))
        return result
    return wrapper


def option():
    parser = argparse.ArgumentParser(
        description=''
    )
    parser.add_argument('-v', '--version',
                        help='version information',
                        action='version',
                        version='%(prog)s 0.1'
                        )
    parser.add_argument('-o', '--output',
                        help='specifiy output folder',
                        action='store',
                        default='out'
                        )
    parser.add_argument('-k', '--keyword-file',
                        help='input keyword file',
                        default='keyword.xlsx')
    parser.add_argument('-a', '--appid_file',
                        help='appid',
                        default='appid.xlsx'
                        )
    parser.add_argument('-m', '--max_number',
                        help='max number of search result',
                        type=int,
                        default=10000)
    return parser.parse_args()

def load_xlsx_cells(xlsx:str):
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[wb.sheetnames[0]]

    values = []
    clm = 0
    for r in range(ws.max_row):
        cell = ws['A'+str(r+1)]
        values.append(cell.value)
    return values


def load_keywords(keyword_file:str):
    keywords = load_xlsx_cells(keyword_file)
    return keywords


def load_appids(appid_file: str):
    t = load_xlsx_cells(appid_file)
    appids = list(map(lambda x: x.strip(), t))
    return appids


def load_limit_conf(keyword_file):
    wb = openpyxl.load_workbook(keyword_file)
    ws = wb[wb.sheetnames[0]]

    max_items_per_xlsx = int(ws['B1'].value)
    max_shops = int(ws['C1'].value)
    return (max_items_per_xlsx, max_shops)


@stop_watch
def main():
    args = option()
    print(args)
    keywords = load_keywords(args.keyword_file)
    appids = load_appids(args.appid_file)
    max_items_per_xlsx, max_shops = load_limit_conf(args.keyword_file)

    os.makedirs(args.output, exist_ok=True)
    searchItem = searchItems(keywords, appids, args.output, args.max_number,
                                max_items_per_xlsx, max_shops)
    searchItem.run()
    return 0


if __name__ == '__main__':
    exit(main())
